import os
from os import path
import shutil
import json
import tkinter
from tkinter import filedialog
import pandas as pd
import json
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import time

root = tkinter.Tk()
root.withdraw()

# Open xlsx file
open_sheet = path.exists("cache/opened_sheet.json")
if open_sheet == True :
  opened_sheet_file_path = "cache/opened_sheet.json"
  json_file = open(opened_sheet_file_path)
  data = json.load(json_file)
  xlsx_sheet_check = path.exists(data ['xlsx_file_path'])
  if xlsx_sheet_check == True :
    xlsx_file_path = data ['xlsx_file_path']
  else :
    shutil.rmtree('cache', ignore_errors=True)
    xlsx_file_path = filedialog.askopenfilename(title="Open Excel-XLSX File")
    cache_path = os.path.join(str(os.getcwd()), "cache")
    dictionary = {"xlsx_file_path" : xlsx_file_path}
    json_object = json.dumps(dictionary, indent = 1)
    with open("cache/opened_sheet.json", "w") as outfile:
      outfile.write(json_object)
else :
  xlsx_file_path = filedialog.askopenfilename(title="Open Excel-XLSX File")
  cache_path = os.path.join(str(os.getcwd()), "cache")
  os.mkdir(cache_path)
  dictionary = {"xlsx_file_path" : xlsx_file_path}
  json_object = json.dumps(dictionary, indent = 1)
  with open("cache/opened_sheet.json", "w") as outfile:
    outfile.write(json_object)

# read imported xlsx file path using pandas
input_workbook = pd.read_excel(xlsx_file_path, sheet_name = 'Sheet1', usecols = 'E:I', dtype=str)
input_workbook.head()

# read total number of rows present in xlsx
number_of_rows = len(input_workbook.index)

# Opening JSON file & returns JSON object as a dictionary
json_file = open('settings.json')
settings_data = json.load(json_file)

input_workbook_cc_number = input_workbook['Card Number'].values.tolist()
input_workbook_cvv_number = input_workbook['CVV'].values.tolist()
input_workbook_expiry_number = input_workbook['Expiry'].values.tolist()
input_workbook_ipin = input_workbook['Ipin'].values.tolist()
input_workbook_desk_number = input_workbook['Desk'].values.tolist()

# get-output sheet to append output
output_sheet = path.exists("Output.xlsx")
if output_sheet == True :
  output_sheet_file_path = "Output.xlsx"
else :
  output_headers= ['FirstName','LastName', 'Mobile', 'Email','Amount', 'CardNumber', 'CVV', 'Expiry', 'Ipin', 'Status', 'Transation No', 'No.of Transactions', 'Desk', "Desk Holder"]
  overall_output = Workbook()
  page = overall_output.active
  page.append(output_headers)
  overall_output.save(filename = 'Output.xlsx')
  output_sheet_file_path = "Output.xlsx"

def cal():
  global output_cc_number
  global done_transactions_wb_1
  global h
  output_load_wb_2 = pd.read_excel(output_sheet_file_path, sheet_name = 'Sheet', usecols = 'F', dtype=str)
  output_load_wb_2.head()
  output_cc_number = output_load_wb_2['CardNumber'].values.tolist()
  output_load_wb_1 = pd.read_excel(output_sheet_file_path, sheet_name = 'Sheet', usecols = 'L', dtype=int)
  output_load_wb_1.head()
  done_transactions_wb_1 = output_load_wb_1['No.of Transactions'].values.tolist()
  h = len(output_load_wb_1.index) - 1
  print (output_cc_number[h],done_transactions_wb_1[h])

def cc_expiry():
  global expiry_month
  global expiry_year
  global expiry_year1
  global expiry_year2
  global expiry_year3
  global expiry_year4
  workbook_expiry_month = input_workbook_expiry_number[x]
  workbook_expiry_year = input_workbook_expiry_number[x]
  expiry_month = workbook_expiry_month[:2]
  expiry_year = workbook_expiry_year[5:]
  expiry_year1 = workbook_expiry_year[3]
  expiry_year2 = workbook_expiry_year[4]
  expiry_year3 = workbook_expiry_year[5]
  expiry_year4 = workbook_expiry_year[6]

def textbox_field(xpath, timeout_time, send_keys_data):
  global timeout_exception
  global timeout_exception1
  try :
    WebDriverWait(driver, timeout=timeout_time).until(ec.visibility_of_element_located((By.XPATH, xpath)))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    timeout_exception1 = False
    textbox_elements = driver.find_element_by_xpath (xpath)
    textbox_elements.send_keys(send_keys_data)

def button_field(button_xpath, timeout_time):
  global timeout_exception
  global timeout_exception1
  try :
    WebDriverWait(driver, timeout=timeout_time).until(ec.visibility_of_element_located((By.XPATH, button_xpath)))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    page_button = driver.find_element_by_xpath (button_xpath)
    page_button.click()


def textbox_field_click(xpath, timeout_time):
  global timeout_exception
  global timeout_exception1
  try :
    WebDriverWait(driver, timeout=timeout_time).until(ec.visibility_of_element_located((By.XPATH, xpath)))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    timeout_exception1 = False
    act.click(driver.find_element_by_xpath (xpath)).perform()

def textbox_field_click_css_selector(xpath, timeout_time):
  global timeout_exception
  global timeout_exception1
  try :
    WebDriverWait(driver, timeout=timeout_time).until(ec.visibility_of_element_located((By.CSS_SELECTOR, xpath)))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = True
  else :
    timeout_exception = False
    timeout_exception1 = False
    act.click(driver.find_element_by_css_selector (xpath)).perform()
  
def start_link():
  driver.get("https://mfq.manbafinance.com/paymentwebsite")

def page_one():
  textbox_field('//*[@id="txtlanno"]', 8, settings_data['LAN'])
  button_field('//*[@id="next"]', 8)

def page_two():
  time.sleep(0.75)
  textbox_field_click('//*[@id="Other"]', 8)
  textbox_field('//*[@id="txtamount"]', 8, settings_data['payable_amount'])
  button_field('//*[@id="btnPay"]', 8)

def page_three():
  global timeout_exception
  global timeout_exception1
  driver.switch_to.frame(WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.CLASS_NAME, "paymtiframe"))))
  time.sleep(0.50)
  driver.switch_to.frame(WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.CLASS_NAME, "razorpay-checkout-frame"))))
  #driver.switch_to.frame(driver.find_element_by_class_name("razorpay-checkout-frame"))
  textbox_field('//*[@id="contact"]', 8, settings_data['registered_mobile_no'])
  textbox_field('//*[@id="email"]', 8, settings_data['email_id'])
  button_field('//*[@id="footer-cta"]', 8)
  time.sleep(0.50)
  try :
    WebDriverWait(driver, timeout=4).until(ec.visibility_of_element_located((By.CSS_SELECTOR, '#form-common > div.screen.screen-comp.svelte-3j22k8 > div > div > div.home-methods.svelte-1ai009r > div.methods-block.svelte-v8dhx4 > div > button.instrument.slotted-option.svelte-1u727jy > div > div.svelte-1u727jy > div:nth-child(1)')))
  except TimeoutException:
    timeout_exception = True
    timeout_exception1 = False
  else :
    pay_type = driver.find_element_by_css_selector('#form-common > div.screen.screen-comp.svelte-3j22k8 > div > div > div.home-methods.svelte-1ai009r > div.methods-block.svelte-v8dhx4 > div > button.instrument.slotted-option.svelte-1u727jy > div > div.svelte-1u727jy > div:nth-child(1)')
    print (pay_type.text)
    if pay_type.text == "Pay using Card":
      textbox_field_click_css_selector('#form-common > div.screen.screen-comp.svelte-3j22k8 > div > div > div.home-methods.svelte-1ai009r > div.methods-block.svelte-v8dhx4 > div > button.instrument.slotted-option.svelte-1u727jy > div > div.svelte-1u727jy > div:nth-child(1)', 8)
    else :
      textbox_field_click_css_selector('#form-common > div.screen.screen-comp.svelte-3j22k8 > div > div > div.home-methods.svelte-1ai009r > div:nth-child(2) > div > button:nth-child(1) > div > div.svelte-1u727jy > div:nth-child(1)', 8)

  textbox_field('//*[@id="card_number"]', 8, input_workbook_cc_number[x])
  textbox_field('//*[@id="card_expiry"]', 8, expiry_month + expiry_year3 + expiry_year4)
  textbox_field('//*[@id="card_name"]', 8, settings_data['first_name'])
  textbox_field('//*[@id="card_cvv"]', 8, input_workbook_cvv_number[x])
  button_field('//*[@id="footer-cta"]', 8)

def page_four():
  driver.switch_to.window(driver.window_handles[1])
  textbox_field('//*[@id="ipin"]', 8, input_workbook_ipin[x])
  button_field('//*[@id="otpbut"]', 8)

def output():
  global output_status
  global transaction_output_status
  global timeout_exception
  global timeout_exception1  
  try :
    WebDriverWait(driver, timeout=2).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="set"]/div/div/div[2]/div/div[3]/font')))
  except TimeoutException:
    driver.switch_to.window(driver.window_handles[0])
    driver.switch_to.frame(WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.CLASS_NAME, "paymtiframe"))))
    #time.sleep(2)
    #driver.switch_to.frame(WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.CLASS_NAME, "razorpay-checkout-frame"))))
    try :
      WebDriverWait(driver, timeout=3).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="checkout-parent"]/div[2]/div[2]/div')))
    except TimeoutException :
      try:
        WebDriverWait(driver, timeout=1).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="fd-t"]')))
      except TimeoutException:
        timeout_exception = True
        timeout_exception1 = True
      else:
        output_status_element = driver.find_element_by_xpath('//*[@id="fd-t"]')
        output_status = output_status_element.text
        transaction_output_status = '-'
        timeout_exception = False
        timeout_exception1 = False
    else :
      output_status_element = driver.find_element_by_xpath('//*[@id="checkout-parent"]/div[2]/div[2]')
      output_status = output_status_element.text
      transaction_element = driver.find_element_by_xpath('//*[@id="checkout-parent"]/div[2]/div[2]/div')
      transaction_output_status = transaction_element.text
      timeout_exception = False
      timeout_exception1 = False
  else :
    try :
      driver.find_element_by_xpath ('//*[@id="cancel"]')
    except NoSuchElementException:
      time.sleep(1)
      driver.find_element_by_xpath ('//*[@id="cancel"]').click()
    else :
      driver.find_element_by_xpath ('//*[@id="cancel"]').click()
    driver.switch_to.window(driver.window_handles[0])
    driver.switch_to.frame(WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.CLASS_NAME, "paymtiframe"))))
    time.sleep(0.50)
    driver.switch_to.frame(WebDriverWait(driver, timeout=8).until(ec.visibility_of_element_located((By.CLASS_NAME, "razorpay-checkout-frame"))))
    output_status = "Please enter correct IPIN / WEB PIN"
    transaction_output_status = "-"
    time.sleep(1)

def output_save():
  entry_list = [[settings_data['first_name'], settings_data['last_name'], settings_data['registered_mobile_no'], settings_data['email_id'], settings_data['payable_amount'], input_workbook_cc_number[x], input_workbook_ipin[x], input_workbook_cvv_number[x], input_workbook_expiry_number[x], output_status, transaction_output_status, z+1, int(input_workbook_desk_number[x]), settings_data["desk_holder"]]]
  output_wb = load_workbook(output_sheet_file_path)
  page = output_wb.active
  for info in entry_list:
      page.append(info)
  output_wb.save(filename='Output.xlsx')

def whole_work():
    start_link()
    page_one()
    page_two()
    cc_expiry()
    page_three()
    page_four()
    output()

# exception
def retry_1 ():
  start_link()
  page_one()
  page_two()
  cc_expiry()
  page_three()
  page_four()
  output()

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--incognito")
caps = DesiredCapabilities().CHROME
caps["pageLoadStrategy"] = "none"
#caps["pageLoadStrategy"] = "eager"
#caps["pageLoadStrategy"] = "normal"
driver=webdriver.Chrome(chrome_options=chrome_options, desired_capabilities=caps, executable_path="chromedriver.exe")
driver.maximize_window()
act = ActionChains(driver)
try:
  cal()
except IndexError:
  for x in range (0 , number_of_rows):
    for z in range (0, int(settings_data['number_of_time_transactions_per_card'])):
      whole_work()
      if timeout_exception == True:
        retry_1()
        if timeout_exception1 == True:
          output_status = "Null"
          transaction_output_status = "-"
        else:
          output_save()
      else:
        output_save()
else:
  last_txncard =  input_workbook[input_workbook['Card Number'] == output_cc_number[h]].index[0]
  for x in range (last_txncard , number_of_rows):
    for z in range (done_transactions_wb_1[h], int(settings_data['number_of_time_transactions_per_card'])):
      whole_work()
      if timeout_exception == True:
        retry_1()
        if timeout_exception1 == True:
          output_status = "Null"
          transaction_output_status = "-"
        else:
          output_save()
      else:
        output_save()
    done_transactions_wb_1[h] = 0

driver.quit()
